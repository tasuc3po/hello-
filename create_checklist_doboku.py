import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "土木申請チェックリスト"

FILL_P1 = PatternFill("solid", fgColor="FFE0E0")
FILL_P2 = PatternFill("solid", fgColor="FFE5CC")
FILL_P3 = PatternFill("solid", fgColor="FFF2CC")
FILL_P4 = PatternFill("solid", fgColor="E2EFDA")
FILL_P5 = PatternFill("solid", fgColor="DAEEF3")
FILL_P6 = PatternFill("solid", fgColor="EAD1DC")
FILL_P7 = PatternFill("solid", fgColor="F3E5F5")
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SECTION = PatternFill("solid", fgColor="D6E4F0")
FILL_FUTURE = PatternFill("solid", fgColor="F5F5F5")

thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws.merge_cells("A1:H1")
c = ws.cell(row=1, column=1, value="建設業許可申請 チェックリスト【土木のみ・第一段階】　沖縄県・法人・一般建設業（土木）新規")
c.font = Font(bold=True, color="FFFFFF", size=12, name="Yu Gothic UI")
c.fill = FILL_HEADER
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:H2")
c = ws.cell(row=2, column=1, value="提出3部（正・副・控）／A4フラットファイルに綴る／各証明書は発行後3ヶ月以内（残高証明書は1ヶ月以内）　※収入証紙4.5万円（土木1業種）")
c.font = Font(size=9, color="333333", name="Yu Gothic UI")
c.fill = PatternFill("solid", fgColor="D6E4F0")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# Header
headers = ["優先度", "✓", "書類名", "摘要・注意事項", "備考タグ", "入手先", "担当", "完了日"]
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=i+1, value=h)
    c.font = Font(bold=True, color="FFFFFF", size=10, name="Yu Gothic UI")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws.row_dimensions[3].height = 22

def section_row(ws, row, text, fill):
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="1F4E79", size=10, name="Yu Gothic UI")
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border
    ws.row_dimensions[row].height = 18

def data_row(ws, row, priority, name, notes, tag, source, person, fill):
    values = [priority, "□", name, notes, tag, source, person, ""]
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=i+1, value=val)
        c.font = Font(size=9, name="Yu Gothic UI")
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center" if i == 1 else "left")
        c.border = border
        c.fill = fill
    ws.row_dimensions[row].height = 42

row = 4

# P1
section_row(ws, row, "【P1】今すぐ着手（ブロッカー）", FILL_SECTION); row += 1
data_row(ws, row, "★1", "面談予約（沖縄県 技術・建設業課）",
         "TEL: 098-866-2374　※書類提出前に予約必須。面談形式の審査あり。「土木のみ新規申請」と伝える。日程が決まれば全体スケジュールが組める",
         "最初にやること", "電話予約", "社長", FILL_P1); row += 1

# P2
section_row(ws, row, "【P2】早めに着手（時間・手間がかかる）", FILL_SECTION); row += 1
data_row(ws, row, "★2", "工事請負契約書・注文書・請求書等の写し",
         "年3件×6年分＝18件必要。建設工事の請負とわかるもの。社内書類の掘り起こしに時間がかかる",
         "", "自社", "担当者", FILL_P2); row += 1
data_row(ws, row, "★2", "財務諸表（貸借対照表・損益計算書・株主資本等変動計算書・注記表）",
         "第15〜17号の2。直前決算期のもの。表紙の添付が必要。税理士への依頼が必要",
         "", "自社（税理士）", "経理", FILL_P2); row += 1
data_row(ws, row, "★2", "東商企業要覧沖縄県版（写し）または許可証明書（写し）",
         "証明期間分。該当ページ＋表紙をセットで提出",
         "", "自社", "担当者", FILL_P2); row += 1

# P3
section_row(ws, row, "【P3】窓口まとめて取得（1回で済ませる）", FILL_SECTION); row += 1
data_row(ws, row, "★3", "履歴事項全部証明書（商業登記簿謄本）①経管用",
         "原本提出。6年分の役員在籍がわかるもの", "原本提出", "法務局", "社長", FILL_P3); row += 1
data_row(ws, row, "★3", "登記事項証明書（商業登記簿）②申請書類用",
         "原本提出。発行後3ヶ月以内", "原本提出", "法務局", "社長", FILL_P3); row += 1
data_row(ws, row, "★3", "社長の後見等登記事項証明書",
         "原本提出", "原本提出", "法務局", "社長", FILL_P3); row += 1
data_row(ws, row, "★3", "社長の住民票抄本",
         "発行後3ヶ月以内。原本提示＋写しを「控え」に添付", "原本提示＋写し", "市区町村", "社長", FILL_P3); row += 1
data_row(ws, row, "★3", "社長の身分証明書",
         "本籍地の役所で取得。原本提出", "原本提出", "本籍地役所", "社長", FILL_P3); row += 1
data_row(ws, row, "★3", "上原の住民票抄本",
         "発行後3ヶ月以内。原本提示＋写しを「控え」に添付", "原本提示＋写し", "市区町村", "上原", FILL_P3); row += 1
data_row(ws, row, "★3", "納税証明書（法人事業税）",
         "原本提出。県税・直前1期分", "原本提出", "県税事務所", "経理", FILL_P3); row += 1

# P4
section_row(ws, row, "【P4】県HPからDL・自分で作成（面談予約後に着手）", FILL_SECTION); row += 1
data_row(ws, row, "★4", "経営業務管理責任者証明書（様式第7号）＋略歴書",
         "沖縄県HPからDL・作成。証明内容の確認書類（写し）を添付", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "許可申請者等の住所・生年月日等に関する調書（第12号）",
         "役員等全員分（別紙1に記載した役員全員）", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "営業所技術者等証明書（様式第8号）＋営業所技術者一覧表（別紙四）",
         "土木担当分（上原）を作成", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "【新規申請用】チェックシート（沖縄県様式）",
         "沖縄県HPからDL。提出時に該当項目にマーカーまたはチェックを入れて提出", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "表紙（建設業許可申請書）＋建設業許可申請書（第1号）",
         "沖縄県HPからDL・作成", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "工事経歴書（第2号）",
         "実績の有無に関わらず必要。工事実績確認できる契約書等の提示が必要", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "直前3年の各事業年度における工事施工金額（第3号）",
         "実績の有無に関わらず必要", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "使用人数（第4号）",
         "沖縄県HPからDL・作成", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "誓約書（第6号）",
         "本文の削除不可。申請者名は代表者名を記入", "", "県HP", "社長", FILL_P4); row += 1
data_row(ws, row, "★4", "営業所一覧表（別紙二(1)）＋営業の沿革（第20号）",
         "営業所の写真（外観・入口・内部）を添付。令和5年12月から必須", "", "県HP＋自社撮影", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "役員等の一覧表（別紙一）",
         "役員＋相談役・顧問・5%以上出資者も記載", "", "県HP", "担当者", FILL_P4); row += 1
data_row(ws, row, "★4", "所属建設業者団体（第20号の2）・主要取引金融機関名（第20号の3）",
         "沖縄県HPからDL・作成", "", "県HP", "担当者", FILL_P4); row += 1

# P5
section_row(ws, row, "【P5】手元書類・社内確認（比較的すぐ用意できる）", FILL_SECTION); row += 1
data_row(ws, row, "★5", "2級土木施工管理技士 資格証明書（写し）",
         "写し添付＋原本を当日持参（必須）", "写し添付＋原本持参", "自己保管", "上原", FILL_P5); row += 1
data_row(ws, row, "★5", "社長の常勤確認書類（健康保険・厚生年金関係）",
         "健保・厚生年金被保険者標準報酬決定通知書の写し、または資格取得届（確認印あり）等", "原本提示", "会社", "経理", FILL_P5); row += 1
data_row(ws, row, "★5", "上原の常勤確認書類（社会保険関係）",
         "健保・厚生年金被保険者標準報酬決定通知書等。原本提示", "原本提示", "会社", "経理", FILL_P5); row += 1
data_row(ws, row, "★5", "健康保険等の加入状況（第7号の3）＋保険料納入領収証書",
         "健保・厚生年金：直前の保険料領収証書または納入証明書の写し／雇用保険：労働保険概算・確定保険料申告書の控え＋領収済通知書の写し",
         "", "会社", "経理", FILL_P5); row += 1
data_row(ws, row, "★5", "定款（写し）",
         "自社保管のもの", "", "自社", "担当者", FILL_P5); row += 1

# P6
section_row(ws, row, "【P6】最後に取得（申請直前）", FILL_SECTION); row += 1
data_row(ws, row, "★6", "預金残高証明書（500万円以上）",
         "申請受付日前1ヶ月以内に発行。他の書類が全部揃ってから最後に取得する",
         "申請直前1ヶ月以内・原本提示", "取引銀行", "社長", FILL_P6); row += 1
data_row(ws, row, "★6", "沖縄県証紙（収入証紙）4.5万円分",
         "別紙三（収入印紙等貼付欄）に貼付。知事許可・新規・土木1業種分（造園追加時はさらに5万円）",
         "当日必要", "郵便局等", "担当者", FILL_P6); row += 1

# Future plan
row += 1
section_row(ws, row, "【第二段階】造園業種追加申請（2026年2月以降・合格後）", FILL_FUTURE); row += 1

future_items = [
    ("予定", "造園施工管理技士（2級）合格証明書", "2025年11月二次試験→2026年2月以降合格発表後に取得", "合格後", "試験機関", "本人"),
    ("予定", "業種追加申請（造園）", "手数料5万円。土木許可取得後に申請。造園技術者関連書類を追加提出", "合格後", "沖縄県", "担当者"),
]
for item in future_items:
    values = [item[0], "□", item[1], item[2], item[3], item[4], item[5], ""]
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=i+1, value=val)
        c.font = Font(size=9, color="888888", name="Yu Gothic UI", italic=True)
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center" if i == 1 else "left")
        c.border = border
        c.fill = FILL_FUTURE
    ws.row_dimensions[row].height = 36
    row += 1

# Notes
row += 1
ws.merge_cells(f"A{row}:H{row}")
c = ws.cell(row=row, column=1, value="【備考・注意事項】")
c.font = Font(bold=True, color="FFFFFF", size=10, name="Yu Gothic UI")
c.fill = FILL_HEADER
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 18
row += 1

notes = [
    "・提出時は「許可申請書及び添付書類一覧」の様式の順に並べてA4フラットファイルに綴ること",
    "・健康保険証の写しを提出する場合は、保険者番号・被保険者等記号・番号をマスキング（黒塗り）すること",
    "・造園は2025年11月二次試験→合格後に業種追加申請（手数料5万円）で対応。新規より安く済む",
    "・沖縄県 土木建築部 技術・建設業課　TEL: 098-866-2374　〒900-8570 那覇市泉崎1-2-2 行政棟11階（北側）",
]
for note in notes:
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value=note)
    c.font = Font(size=9, name="Yu Gothic UI")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.border = border
    ws.row_dimensions[row].height = 18
    row += 1

# Column widths
for i, w in enumerate([12, 4, 35, 55, 20, 14, 12, 12]):
    ws.column_dimensions[get_column_letter(i+1)].width = w

ws.freeze_panes = "A4"

# Legend sheet
ws2 = wb.create_sheet("優先度の見方")
legend = [
    ("優先度", "意味", "目安", ""),
    ("★1 今すぐ（赤）", "ここから全体が動き出す", "今日中", FILL_P1),
    ("★2 早めに着手（オレンジ）", "時間・手間がかかる書類", "今週中に着手", FILL_P2),
    ("★3 窓口まとめて（黄）", "法務局・市区町村・県税事務所を1回で", "面談予約後に日程調整", FILL_P3),
    ("★4 県HPからDL・作成（緑）", "様式をDLして自分で作成", "面談予約後に作成開始", FILL_P4),
    ("★5 手元書類・社内確認（青）", "社内・手元にある書類の確認", "並行して進める", FILL_P5),
    ("★6 最後に取得（ピンク）", "有効期限が短いため申請直前に", "全書類が揃ってから", FILL_P6),
    ("予定 造園追加（グレー）", "2026年2月以降・合格後に対応", "第二段階", FILL_FUTURE),
]
for i, row_data in enumerate(legend):
    fill = row_data[3] if len(row_data) > 3 and row_data[3] else PatternFill("solid", fgColor="2E75B6")
    fc = "FFFFFF" if i == 0 else "000000"
    for j, val in enumerate(row_data[:3]):
        c = ws2.cell(row=i+1, column=j+1, value=val)
        c.font = Font(bold=(i==0), color=fc, size=10, name="Yu Gothic UI")
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = border
    ws2.row_dimensions[i+1].height = 28
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 40
ws2.column_dimensions["C"].width = 25

output = "/home/user/hello-/建設業許可チェックリスト_土木先行申請版.xlsx"
wb.save(output)
print(f"保存完了: {output}")
