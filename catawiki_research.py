#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Catawiki 落札価格リサーチツール v1.0
個人リサーチ・仕入れ判断目的専用
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import time
import re
from datetime import datetime
import json
import os
import threading

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9,ja;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
}


class CatawikiResearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Catawiki 落札価格リサーチツール v1.0")
        self.root.geometry("980x720")
        self.root.configure(bg='#f0f2f5')
        self.root.resizable(True, True)

        self.results = []
        self.is_searching = False

        self._build_ui()

    # ------------------------------------------------------------------ UI --
    def _build_ui(self):
        # ヘッダーバー
        header = tk.Frame(self.root, bg='#0096C7', height=56)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="🔍  Catawiki 落札価格リサーチ",
                 bg='#0096C7', fg='white',
                 font=('Hiragino Sans', 15, 'bold')).pack(side='left', padx=20, pady=12)
        tk.Label(header, text="個人リサーチ専用 ／ 礼儀正しい低速アクセス",
                 bg='#0096C7', fg='#cce9f5',
                 font=('Hiragino Sans', 9)).pack(side='right', padx=20)

        # アクセントライン
        tk.Frame(self.root, bg='#FF6B35', height=3).pack(fill='x')

        # 検索エリア
        sf = tk.Frame(self.root, bg='#f0f2f5', pady=14)
        sf.pack(fill='x', padx=20)

        tk.Label(sf, text="キーワード:", bg='#f0f2f5',
                 font=('Hiragino Sans', 11)).grid(row=0, column=0, sticky='w')

        self.keyword_var = tk.StringVar()
        kw_entry = tk.Entry(sf, textvariable=self.keyword_var,
                            font=('Hiragino Sans', 12), width=32,
                            relief='flat', bd=2, highlightthickness=2,
                            highlightbackground='#ccc', highlightcolor='#0096C7')
        kw_entry.grid(row=0, column=1, padx=10, ipady=5)
        kw_entry.bind('<Return>', lambda e: self.start_search())

        tk.Label(sf, text="例: Georg Jensen / Seiko / G-SHOCK",
                 bg='#f0f2f5', fg='#888',
                 font=('Hiragino Sans', 9)).grid(row=1, column=1, sticky='w', padx=10)

        tk.Label(sf, text="取得ページ数:", bg='#f0f2f5',
                 font=('Hiragino Sans', 11)).grid(row=0, column=2, padx=(20, 5))
        self.pages_var = tk.StringVar(value='3')
        tk.Spinbox(sf, from_=1, to=10, textvariable=self.pages_var,
                   width=4, font=('Hiragino Sans', 11)).grid(row=0, column=3)
        tk.Label(sf, text="ページ（1ページ≒20件）",
                 bg='#f0f2f5', fg='#888',
                 font=('Hiragino Sans', 9)).grid(row=1, column=3, columnspan=2, sticky='w')

        self.search_btn = tk.Button(sf, text="▶  リサーチ開始",
                                    command=self.start_search,
                                    bg='#FF6B35', fg='white',
                                    font=('Hiragino Sans', 12, 'bold'),
                                    relief='flat', padx=22, pady=7,
                                    cursor='hand2', activebackground='#e55a24')
        self.search_btn.grid(row=0, column=5, rowspan=2, padx=20)

        # セパレーター
        ttk.Separator(self.root, orient='horizontal').pack(fill='x', padx=20)

        # 結果テーブル
        tf = tk.Frame(self.root, bg='#f0f2f5')
        tf.pack(fill='both', expand=True, padx=20, pady=8)

        self.count_label = tk.Label(tf, text="結果: 0 件",
                                    bg='#f0f2f5', fg='#333',
                                    font=('Hiragino Sans', 10, 'bold'))
        self.count_label.pack(anchor='w', pady=(0, 4))

        cols = ('title', 'price', 'status', 'date', 'category')
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Research.Treeview.Heading',
                        background='#0096C7', foreground='white',
                        font=('Hiragino Sans', 10, 'bold'))
        style.configure('Research.Treeview', rowheight=24, font=('Hiragino Sans', 10))
        style.map('Research.Treeview', background=[('selected', '#cce9f5')])

        tree_frame = tk.Frame(tf, bg='#f0f2f5')
        tree_frame.pack(fill='both', expand=True)

        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings',
                                  height=16, style='Research.Treeview')
        self.tree.heading('title',    text='タイトル',    anchor='w')
        self.tree.heading('price',    text='落札価格',    anchor='e')
        self.tree.heading('status',   text='状態',       anchor='center')
        self.tree.heading('date',     text='終了日',      anchor='center')
        self.tree.heading('category', text='カテゴリ',    anchor='w')

        self.tree.column('title',    width=380, anchor='w',      stretch=True)
        self.tree.column('price',    width=100, anchor='e',      stretch=False)
        self.tree.column('status',   width=90,  anchor='center', stretch=False)
        self.tree.column('date',     width=100, anchor='center', stretch=False)
        self.tree.column('category', width=180, anchor='w',      stretch=True)

        sb_y = ttk.Scrollbar(tree_frame, orient='vertical',   command=self.tree.yview)
        sb_x = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        sb_y.grid(row=0, column=1, sticky='ns')
        sb_x.grid(row=1, column=0, sticky='ew')
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # ダブルクリックでURL表示
        self.tree.bind('<Double-1>', self._on_row_double_click)

        # ボタン列
        bf = tk.Frame(self.root, bg='#f0f2f5', pady=6)
        bf.pack(fill='x', padx=20)

        tk.Button(bf, text="📊  Excelで保存",
                  command=self.save_to_excel,
                  bg='#217346', fg='white',
                  font=('Hiragino Sans', 11, 'bold'),
                  relief='flat', padx=16, pady=5,
                  cursor='hand2').pack(side='right', padx=4)

        tk.Button(bf, text="🗑  クリア",
                  command=self.clear_results,
                  bg='#888', fg='white',
                  font=('Hiragino Sans', 11),
                  relief='flat', padx=16, pady=5,
                  cursor='hand2').pack(side='right', padx=4)

        self.avg_label = tk.Label(bf, text="",
                                   bg='#f0f2f5', fg='#0096C7',
                                   font=('Hiragino Sans', 10, 'bold'))
        self.avg_label.pack(side='left')

        # ログエリア
        lf = tk.Frame(self.root, bg='#f0f2f5')
        lf.pack(fill='x', padx=20, pady=(0, 12))

        tk.Label(lf, text="ログ", bg='#f0f2f5',
                 font=('Hiragino Sans', 9), fg='#666').pack(anchor='w')
        self.log_text = scrolledtext.ScrolledText(lf, height=5,
                                                   font=('Courier', 9),
                                                   bg='#1e1e1e', fg='#4ec9b0',
                                                   insertbackground='white',
                                                   state='disabled')
        self.log_text.pack(fill='x')

    # --------------------------------------------------------------- Actions --
    def log(self, message):
        self.log_text.configure(state='normal')
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_text.insert('end', f"[{ts}] {message}\n")
        self.log_text.see('end')
        self.log_text.configure(state='disabled')
        self.root.update_idletasks()

    def start_search(self):
        keyword = self.keyword_var.get().strip()
        if not keyword:
            messagebox.showwarning("入力エラー", "キーワードを入力してください")
            return
        if self.is_searching:
            return
        self.is_searching = True
        self.search_btn.configure(state='disabled', text="検索中…", bg='#aaa')
        self.log(f"キーワード「{keyword}」で検索開始")
        threading.Thread(target=self._search_thread,
                         args=(keyword,), daemon=True).start()

    def _search_thread(self, keyword):
        try:
            max_pages = int(self.pages_var.get())
            new_results = self._fetch_catawiki(keyword, max_pages)
            self.root.after(0, lambda: self._update_table(new_results))
        except Exception as e:
            self.root.after(0, lambda: self.log(f"予期せぬエラー: {e}"))
        finally:
            self.root.after(0, self._search_done)

    # --------------------------------------------------------------- Scrape --
    def _fetch_catawiki(self, keyword, max_pages):
        results = []
        session = requests.Session()
        session.headers.update(HEADERS)

        for page in range(1, max_pages + 1):
            encoded = requests.utils.quote(keyword)
            url = f"https://www.catawiki.com/en/search?q={encoded}&page={page}"
            self.root.after(0, lambda p=page: self.log(f"ページ {p} 取得中…"))

            try:
                resp = session.get(url, timeout=20)
                resp.raise_for_status()
                soup = BeautifulSoup(resp.text, 'html.parser')

                # ① __NEXT_DATA__ JSON を優先取得
                tag = soup.find('script', {'id': '__NEXT_DATA__'})
                page_lots = []
                if tag and tag.string:
                    try:
                        nd = json.loads(tag.string)
                        page_lots = self._parse_next_data(nd)
                        self.root.after(0, lambda n=len(page_lots), p=page:
                                        self.log(f"  → ページ {p}: {n} 件（JSONモード）"))
                    except Exception:
                        pass

                # ② JSONが空なら HTML 直接パース
                if not page_lots:
                    page_lots = self._parse_html(soup)
                    self.root.after(0, lambda n=len(page_lots), p=page:
                                    self.log(f"  → ページ {p}: {n} 件（HTMLモード）"))

                results.extend(page_lots)

                if page < max_pages:
                    time.sleep(2.5)   # サーバー負荷を抑える

            except requests.Timeout:
                self.root.after(0, lambda p=page: self.log(f"タイムアウト（ページ {p}）"))
                break
            except requests.HTTPError as e:
                self.root.after(0, lambda e=e: self.log(f"HTTP エラー: {e}"))
                break
            except Exception as e:
                self.root.after(0, lambda e=e: self.log(f"エラー: {e}"))
                break

        return results

    def _parse_next_data(self, data):
        lots = []

        def walk(obj, depth=0):
            if depth > 18 or not obj:
                return
            if isinstance(obj, dict):
                tp = obj.get('__typename', '')
                has_title = isinstance(obj.get('title'), str) and len(obj['title']) > 3
                has_id    = isinstance(obj.get('id'), (int, str))
                has_price = any(k in obj for k in
                                ('closingPrice', 'soldPrice', 'currentBid',
                                 'buyNowPrice', 'estimatedValue'))
                if has_title and has_id and (has_price or 'closingDate' in obj or 'Lot' in tp):
                    lot = self._extract_fields(obj)
                    if lot:
                        lots.append(lot)
                for v in obj.values():
                    walk(v, depth + 1)
            elif isinstance(obj, list):
                for item in obj:
                    walk(item, depth + 1)

        walk(data)

        # 重複除去
        seen, unique = set(), []
        for lot in lots:
            key = lot['title'] + str(lot['price'])
            if key not in seen:
                seen.add(key)
                unique.append(lot)
        return unique

    def _extract_fields(self, obj):
        lot = {}
        lot['title'] = obj.get('title', obj.get('name', '')).strip()
        if not lot['title']:
            return None

        # 価格
        for field in ('closingPrice', 'soldPrice', 'buyNowPrice',
                      'currentBid', 'estimatedValue'):
            raw = obj.get(field)
            if raw:
                if isinstance(raw, dict):
                    lot['price']    = raw.get('amount', raw.get('value', ''))
                    lot['currency'] = raw.get('currency', 'EUR')
                elif isinstance(raw, (int, float)):
                    lot['price']    = raw
                    lot['currency'] = 'EUR'
                else:
                    lot['price']    = raw
                    lot['currency'] = 'EUR'
                break
        else:
            lot['price']    = ''
            lot['currency'] = 'EUR'

        # 状態
        raw_status = str(obj.get('status', obj.get('state', ''))).lower()
        status_map = {
            'closed': '終了', 'sold': '落札済み', 'live': '進行中',
            'upcoming': '予定', 'unsold': '不落', 'published': '公開中',
        }
        lot['status'] = status_map.get(raw_status, raw_status)

        # 日付
        d = obj.get('closingDate', obj.get('endDate', obj.get('closedAt', '')))
        lot['date'] = str(d)[:10] if d and 'T' in str(d) else str(d or '')

        # カテゴリ
        cat = obj.get('category', obj.get('categoryName', {}))
        lot['category'] = cat.get('name', '') if isinstance(cat, dict) else str(cat or '')

        # URL
        lot_id = obj.get('id', '')
        lot['url'] = f"https://www.catawiki.com/en/lots/{lot_id}" if lot_id else ''

        return lot

    def _parse_html(self, soup):
        """フォールバック：HTML 直接パース"""
        lots = []
        selectors = [
            'article[data-testid]', '.lot-card', '[class*="lot-item"]',
            '[class*="search-result"]', 'li[class*="lot"]',
        ]
        for sel in selectors:
            items = soup.select(sel)
            if not items:
                continue
            for item in items:
                lot = {}
                te = item.find(['h2', 'h3', 'h4']) or item.find(class_=re.compile('title'))
                lot['title'] = te.get_text(strip=True) if te else ''
                pm = re.search(r'€\s*([\d,.]+)', item.get_text())
                lot['price']    = pm.group(1).replace(',', '') if pm else ''
                lot['currency'] = 'EUR'
                lot['status']   = ''
                lot['date']     = ''
                lot['category'] = ''
                a = item.find('a', href=True)
                href = a['href'] if a else ''
                lot['url'] = f"https://www.catawiki.com{href}" if href.startswith('/') else href
                if lot['title']:
                    lots.append(lot)
            break
        return lots

    # ----------------------------------------------------------- UI updates --
    def _update_table(self, new_results):
        self.results.extend(new_results)
        self._redraw_tree()
        self._update_stats()
        self.log(f"完了 ─ 累計 {len(self.results)} 件")

    def _redraw_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, lot in enumerate(self.results):
            price = lot.get('price', '')
            try:
                price_str = f"€ {float(str(price).replace(',', '')):.2f}"
            except (ValueError, TypeError):
                price_str = f"€ {price}" if price else '─'
            tag = 'even' if i % 2 == 0 else 'odd'
            self.tree.insert('', 'end', iid=str(i),
                             values=(lot.get('title', ''),
                                     price_str,
                                     lot.get('status', ''),
                                     lot.get('date', ''),
                                     lot.get('category', '')),
                             tags=(tag,))
        self.tree.tag_configure('even', background='#ffffff')
        self.tree.tag_configure('odd',  background='#f0f9ff')
        self.count_label.configure(text=f"結果: {len(self.results)} 件")

    def _update_stats(self):
        prices = []
        for lot in self.results:
            try:
                prices.append(float(str(lot.get('price', '')).replace(',', '')))
            except (ValueError, TypeError):
                pass
        if prices:
            avg = sum(prices) / len(prices)
            mx  = max(prices)
            mn  = min(prices)
            self.avg_label.configure(
                text=f"平均: €{avg:.0f}  最高: €{mx:.0f}  最低: €{mn:.0f}  有効: {len(prices)}件")
        else:
            self.avg_label.configure(text='')

    def _search_done(self):
        self.is_searching = False
        self.search_btn.configure(state='normal', text="▶  リサーチ開始", bg='#FF6B35')

    def clear_results(self):
        self.results = []
        self._redraw_tree()
        self.avg_label.configure(text='')
        self.log("クリアしました")

    def _on_row_double_click(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        lot = self.results[idx]
        url = lot.get('url', '')
        title = lot.get('title', '')
        if url:
            import webbrowser
            webbrowser.open(url)
            self.log(f"ブラウザで開きました: {title[:40]}")
        else:
            messagebox.showinfo("URL なし", "このロットのURLは取得できませんでした")

    # --------------------------------------------------------------- Excel --
    def save_to_excel(self):
        if not self.results:
            messagebox.showwarning("データなし", "保存するデータがありません")
            return

        now_str  = datetime.now().strftime('%Y%m%d_%H%M%S')
        kw       = re.sub(r'[\\/:*?"<>|]', '_', self.keyword_var.get().strip())[:20]
        filename = f"catawiki_{kw}_{now_str}.xlsx"
        desktop  = os.path.join(os.path.expanduser("~"), "Desktop")
        filepath = os.path.join(desktop, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catawiki落札データ"

        # ヘッダー
        header_row = ['タイトル', '落札価格(EUR)', '状態', '終了日', 'カテゴリ', 'URL']
        fill_hdr = PatternFill(start_color='0096C7', end_color='0096C7', fill_type='solid')
        font_hdr = Font(color='FFFFFF', bold=True, size=11)
        for c, h in enumerate(header_row, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = fill_hdr
            cell.font = font_hdr
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 22

        # データ
        fill_even = PatternFill(start_color='F0F9FF', end_color='F0F9FF', fill_type='solid')
        for ri, lot in enumerate(self.results, 2):
            ws.cell(row=ri, column=1, value=lot.get('title', ''))
            price = lot.get('price', '')
            try:
                ws.cell(row=ri, column=2, value=float(str(price).replace(',', '')))
                ws.cell(row=ri, column=2).number_format = '0.00'
            except (ValueError, TypeError):
                ws.cell(row=ri, column=2, value=price)
            ws.cell(row=ri, column=3, value=lot.get('status', ''))
            ws.cell(row=ri, column=4, value=lot.get('date', ''))
            ws.cell(row=ri, column=5, value=lot.get('category', ''))
            url = lot.get('url', '')
            if url:
                c6 = ws.cell(row=ri, column=6, value='開く')
                c6.hyperlink = url
                c6.font = Font(color='0563C1', underline='single')
                c6.alignment = Alignment(horizontal='center')

            if ri % 2 == 0:
                for c in range(1, 7):
                    ws.cell(row=ri, column=c).fill = fill_even

        # 列幅
        ws.column_dimensions['A'].width = 52
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 28
        ws.column_dimensions['F'].width = 8

        ws.auto_filter.ref  = f"A1:F{len(self.results) + 1}"
        ws.freeze_panes     = 'A2'

        # 統計シート
        ws2 = wb.create_sheet("統計")
        prices = []
        for lot in self.results:
            try:
                prices.append(float(str(lot.get('price', '')).replace(',', '')))
            except (ValueError, TypeError):
                pass

        ws2['A1'] = 'キーワード'
        ws2['B1'] = self.keyword_var.get().strip()
        ws2['A2'] = '取得件数'
        ws2['B2'] = len(self.results)
        ws2['A3'] = '価格あり件数'
        ws2['B3'] = len(prices)
        if prices:
            ws2['A4'] = '平均落札価格 (€)'
            ws2['B4'] = round(sum(prices) / len(prices), 2)
            ws2['A5'] = '最高落札価格 (€)'
            ws2['B5'] = max(prices)
            ws2['A6'] = '最低落札価格 (€)'
            ws2['B6'] = min(prices)
        ws2['A8'] = 'リサーチ日時'
        ws2['B8'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for row in ws2.iter_rows(min_row=1, max_row=8, min_col=1, max_col=2):
            for cell in row:
                if cell.column == 1:
                    cell.font = Font(bold=True, color='0096C7')

        try:
            wb.save(filepath)
            messagebox.showinfo("保存完了",
                                f"デスクトップに保存しました！\n\nファイル名: {filename}\n件数: {len(self.results)} 件")
            self.log(f"Excel保存完了: {filepath}")
        except PermissionError:
            messagebox.showerror("保存エラー",
                                 "ファイルが開かれています。Excelを閉じてから再試行してください。")
        except Exception as e:
            messagebox.showerror("保存エラー", f"保存に失敗しました:\n{e}")


# ----------------------------------------------------------------- main -----
if __name__ == '__main__':
    root = tk.Tk()
    app  = CatawikiResearchApp(root)
    root.mainloop()
